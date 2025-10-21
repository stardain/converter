"""
Processing file. All processing is done by specialized libraries.
"""

import csv
from PIL import Image
import ffmpeg
from pdf2docx import Converter
from docx2pdf import convert
from openpyxl import Workbook, load_workbook

def process(picked_format, file):

    """
    Function containing all possible convertations. 
    1. Receives a file from controller,
    2. converts it,
    3. saves it here and now.
    """

    if picked_format == "csv": 

        try:
            wb = Workbook()
            worksheet = wb.active

            with open(file, "r", newline="", encoding="utf-8") as input_file:
                reader = csv.reader(input_file, delimiter=',')
                for row in reader:
                    worksheet.append(row)
            wb.save(f"{file.rsplit('.')[0]}.xlsx")

        except Exception:
            print("CONVERTING error (from csv to xlsx).")
        else:
            print("Successful csv to xlsx convertation.")


    if picked_format == "xlsx": 

        try:
            wb = load_workbook(file)
            worksheet = wb.active

            with open(f"{file.rsplit('.')[0]}.csv", 'w', newline="", encoding="utf-8") as output:
                csv_writer = csv.writer(output)
                for row in worksheet.iter_rows():
                    csv_writer.writerow([cell.value for cell in row])

        except Exception:
            print("CONVERTING error (xlsx to csv).")
        else:
            print("Successful csv to xlsx convertation.")


    if picked_format == "jpg": 

        try:
            original_image = Image.open(file)
            final_image = f"{file.rsplit('.')[0]}.png"
            original_image.save(final_image)
        except Exception:
            print("CONVERTING error (jpg to png).")
        else:
            print("Successful jpg to png convertation.")


    if picked_format == "mp4":

        output_audio_path = f'{file.rsplit('.')[0]}.mp3'

        try:
            ffmpeg.input(file).output(output_audio_path, acodec='libmp3lame').run(overwrite_output=True)
        except ffmpeg.Error as e:
            print(f"Error made by converting library: {e}.")
        else:
            print("Successful mp4 to mp3 convertation.")


    if picked_format == "docx": 

        converted_file = f"{file.rsplit('.')[0]}.pdf"

        try:
            convert(file, converted_file)
        except Exception as e:
            print(f"CONVERTING error: {e}")
        else:
            print("Successful docx to pdf convertion.")


    if picked_format == "pdf": 

        converted_file = f"{file.rsplit('.')[0]}.docx"

        try:
            cv = Converter(file)
            cv.convert(converted_file)
            cv.close()
        except Exception:
            print("CONVERTING error (pdf to docx).")
        else:
            print("Successful pdf to docx converting.")
